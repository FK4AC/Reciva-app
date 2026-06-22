#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EDITOR DE VOLCADO INGESAM -> AIR-E (convenio 2087)
===================================================
Permite editar la informacion de la facturacion mensual (subcategoria de
cada usuario, tarifas, altas/bajas) y exportar los 5 archivos respetando
EXACTAMENTE el formato de ancho fijo que exige AIR-E:

  - INGESAM_VOLCADO_2087_AAAAMM.txt        (lineas FC de 235 caracteres)
  - INFO_ADICIONAL_INGESAM_2087_AAAAMM.txt (lineas FD de 2000 caracteres)
  - Hoja1_INGESAM_VOLCADO_2087_AAAAMM.txt
  - Hoja1_INFO_ADICIONAL_INGESAM_2087_AAAAMM.txt
  - TABLA_PRECIOS_ASEO.xlsx

Reglas del formato (verificadas contra archivos reales):
  FC: 'FC' + id(suscriptor+'1261') alineado a la DERECHA terminando en col 16
      + valor en centavos sin separadores + 'DB' terminando en col 45
      + espacios + periodo AAAAMM en cols 229-234. Largo total: 235.
  FD: 'FD' + suscriptor alineado a la DERECHA terminando en col 12
      + espacios + XML alineado a la DERECHA terminando en col 2000.
  El XML se edita de forma "quirurgica" (reemplazo de etiquetas sobre el
  original) para conservar todas sus particularidades (ESTRATO 'o',
  TAFNA/VBA con punto decimal, MES_1 triplicado, etc.).

Uso:  python editor_volcado.py [carpeta_con_archivos]
Requiere: openpyxl  (pip install openpyxl)
"""

import os
import re
import sys
import shutil
from collections import OrderedDict, Counter

SUFIJO_SERVICIO = "1261"      # sufijo del ID en el archivo VOLCADO
CONVENIO = "2087"
LARGO_FC = 235
LARGO_FD = 2000
FIN_ID_FC = 16                # el ID termina (exclusive) en esta columna
FIN_VALOR_FC = 45             # el valor+DB termina en esta columna
COL_PERIODO_FC = 229          # el periodo inicia en esta columna
FIN_ID_FD = 12                # el suscriptor del FD termina en esta columna

MESES = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO",
         "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]

# Etiquetas que se copian del "donante" cuando un usuario cambia de subcategoria
TAGS_TARIFA = ["USO", "TARIFA_MEDIA", "TRLU", "TRBL", "TRRA", "TRA", "TRNA",
               "TAFA", "TAFNA", "VBA", "TC", "TLU", "TBL", "TRT", "TDF",
               "TTL", "TA", "TOTAL"]


# ----------------------------------------------------------------------
# utilidades de numeros y XML
# ----------------------------------------------------------------------
def a_numero(texto):
    """'17.720,00' -> 17720.0   |   '+14.763,55' -> 14763.55"""
    t = texto.strip().replace(".", "").replace(",", ".")
    return float(t) if t not in ("", "+", "-") else 0.0


def a_texto(valor, signo=False):
    """17720.0 -> '17.720,00'  (formato colombiano)"""
    s = f"{abs(valor):,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
    if signo:
        s = ("-" if valor < 0 else "+") + s
    elif valor < 0:
        s = "-" + s
    return s


def tag(xml, nombre):
    m = re.search(rf"<{nombre}>([^<]*)</{nombre}>", xml)
    return m.group(1) if m else None


def poner_tag(xml, nombre, valor):
    """Reemplaza el contenido de la PRIMERA aparicion de la etiqueta."""
    return re.sub(rf"(<{nombre}>)[^<]*(</{nombre}>)",
                  lambda m: m.group(1) + valor + m.group(2), xml, count=1)


def poner_subsidio(xml, valor_txt, pct_txt):
    """SUBSIDIO_CONTRIBUCION tiene VALOR y PORCENTAJE anidados."""
    bloque = (f"<SUBSIDIO_CONTRIBUCION><VALOR>{valor_txt}</VALOR>"
              f"<PORCENTAJE>{pct_txt}</PORCENTAJE></SUBSIDIO_CONTRIBUCION>")
    return re.sub(r"<SUBSIDIO_CONTRIBUCION>.*?</SUBSIDIO_CONTRIBUCION>",
                  bloque, xml, count=1)


# ----------------------------------------------------------------------
# construccion de lineas con el formato exacto
# ----------------------------------------------------------------------
def linea_fc(suscriptor, centavos, periodo):
    id_completo = f"{suscriptor}{SUFIJO_SERVICIO}"
    campo_valor = f"{centavos}DB"
    linea = "FC" + id_completo.rjust(FIN_ID_FC - 2)
    linea = linea + campo_valor.rjust(FIN_VALOR_FC - len(linea))
    linea = linea.ljust(COL_PERIODO_FC) + periodo
    if len(linea) != LARGO_FC:
        raise ValueError(f"Linea FC de {len(linea)} caracteres (suscriptor {suscriptor})")
    return linea


def linea_fd(suscriptor, xml):
    linea = "FD" + suscriptor.rjust(FIN_ID_FD - 2)
    relleno = LARGO_FD - len(linea) - len(xml)
    if relleno < 0:
        raise ValueError(f"XML demasiado largo para el suscriptor {suscriptor}")
    return linea + " " * relleno + xml


# ----------------------------------------------------------------------
# modelo
# ----------------------------------------------------------------------
class Registro:
    """Una factura: su linea FC y su linea FD (XML)."""

    def __init__(self, suscriptor, centavos, periodo, xml, lote):
        self.suscriptor = suscriptor
        self.centavos = centavos
        self.periodo = periodo
        self.xml = xml
        self.lote = lote          # 'principal' o 'hoja1'

    @property
    def uso(self):
        return tag(self.xml, "USO")

    @property
    def total(self):
        return a_numero(tag(self.xml, "TOTAL"))

    def fc(self):
        return linea_fc(self.suscriptor, self.centavos, self.periodo)

    def fd(self):
        return linea_fd(self.suscriptor, self.xml)


class Volcado:
    def __init__(self, carpeta, periodo_archivo):
        self.carpeta = carpeta
        self.periodo_archivo = periodo_archivo      # AAAAMM del NOMBRE de archivo
        self.registros = OrderedDict()              # (lote, suscriptor) -> Registro
        self.cambios = []
        self._cargar()

    # ---------------- carga ----------------
    def _ruta(self, plantilla):
        return os.path.join(self.carpeta, plantilla.format(p=self.periodo_archivo))

    def _cargar(self):
        pares = [("principal", "INGESAM_VOLCADO_{p}.txt".replace("{p}", f"{CONVENIO}_{{p}}"),
                  "INFO_ADICIONAL_INGESAM_{p}.txt".replace("{p}", f"{CONVENIO}_{{p}}")),
                 ("hoja1", "Hoja1_INGESAM_VOLCADO_" + CONVENIO + "_{p}.txt",
                  "Hoja1_INFO_ADICIONAL_INGESAM_" + CONVENIO + "_{p}.txt")]
        for lote, f_vol, f_info in pares:
            vol, info = self._ruta(f_vol), self._ruta(f_info)
            if not (os.path.exists(vol) and os.path.exists(info)):
                print(f"  AVISO: no se encontro el lote '{lote}' ({vol})")
                continue
            fcs = OrderedDict()
            for ln in open(vol, encoding="ascii"):
                ln = ln.rstrip("\n")
                m = re.match(r"FC\s+(\d+)\s", ln)
                id_completo = m.group(1)
                if not id_completo.endswith(SUFIJO_SERVICIO):
                    raise ValueError(f"ID sin sufijo {SUFIJO_SERVICIO}: {id_completo}")
                susc = id_completo[:-len(SUFIJO_SERVICIO)]
                centavos = int(re.search(r"(\d+)DB", ln).group(1))
                periodo = ln[COL_PERIODO_FC:COL_PERIODO_FC + 6]
                fcs[susc] = (centavos, periodo)
            for ln in open(info, encoding="ascii"):
                ln = ln.rstrip("\n")
                susc = ln[2:FIN_ID_FD].strip()
                xml = ln[ln.find("<INFO_ASEO>"):]
                if susc not in fcs:
                    raise ValueError(f"FD sin FC: {susc} en lote {lote}")
                centavos, periodo = fcs[susc]
                self.registros[(lote, susc)] = Registro(susc, centavos, periodo, xml, lote)
            print(f"  Lote {lote}: {len(fcs)} facturas cargadas.")

    # ---------------- tarifario aprendido de los propios archivos ----------------
    def tarifario(self):
        """{uso: (centavos, registro_donante)} aprendido de los registros."""
        t = {}
        for r in self.registros.values():
            t.setdefault(r.uso, (r.centavos, r))
        return dict(sorted(t.items()))

    # ---------------- operaciones de edicion ----------------
    def buscar(self, suscriptor):
        return [r for (l, s), r in self.registros.items() if s == suscriptor]

    def cambiar_uso(self, suscriptor, uso_nuevo):
        """Cambia la subcategoria de un usuario copiando del donante:
        desglose tarifario, subsidio/contribucion, tarifa media y total.
        Conserva su historico y su periodo."""
        regs = self.buscar(suscriptor)
        if not regs:
            raise ValueError("Suscriptor no encontrado.")
        donante = self.tarifario().get(uso_nuevo)
        if donante is None:
            raise ValueError(f"No hay ningun registro con uso '{uso_nuevo}' del cual copiar la tarifa.")
        centavos, don = donante
        for r in regs:
            anterior = r.uso
            for t in TAGS_TARIFA:
                valor_don = tag(don.xml, t)
                if valor_don is not None:
                    r.xml = poner_tag(r.xml, t, valor_don)
            r.xml = poner_subsidio(r.xml,
                                   tag(don.xml, "VALOR"), tag(don.xml, "PORCENTAJE"))
            r.centavos = centavos
            self.cambios.append(f"{suscriptor} ({r.lote}): {anterior} -> {uso_nuevo}")

    def cambiar_tarifa(self, uso, total_nuevo):
        """Cambia la tarifa de una subcategoria completa. Escala el desglose
        (TC, TBL, TRT, TDF, TTL, TLU...) y el subsidio proporcionalmente para
        que sigan cuadrando con el nuevo TOTAL."""
        afectados = [r for r in self.registros.values() if r.uso == uso]
        if not afectados:
            raise ValueError(f"No hay registros con uso '{uso}'.")
        total_viejo = afectados[0].total
        factor = total_nuevo / total_viejo
        componentes = ["TC", "TLU", "TBL", "TRT", "TDF", "TTL", "TA",
                       "TRLU", "TRBL", "TRRA", "TRA", "TRNA", "TAFA", "TARIFA_MEDIA"]
        for r in afectados:
            for t in componentes:
                v = tag(r.xml, t)
                if v is not None and a_numero(v) != 0:
                    r.xml = poner_tag(r.xml, t, a_texto(a_numero(v) * factor))
            v_sub = tag(r.xml, "VALOR")
            if v_sub is not None and a_numero(v_sub) != 0:
                r.xml = poner_subsidio(r.xml, a_texto(a_numero(v_sub) * factor, signo=True),
                                       tag(r.xml, "PORCENTAJE"))
            r.xml = poner_tag(r.xml, "TOTAL", a_texto(total_nuevo))
            r.centavos = int(round(total_nuevo * 100))
        self.cambios.append(f"Tarifa '{uso}': {a_texto(total_viejo)} -> {a_texto(total_nuevo)} ({len(afectados)} usuarios)")

    def alta(self, suscriptor, uso, lote="principal"):
        """Crea un usuario nuevo clonando la estructura de un donante de la
        misma subcategoria, con historico de facturacion en cero."""
        if not suscriptor.isdigit() or not 1 <= len(suscriptor) <= 7:
            raise ValueError("El codigo debe ser numerico, de maximo 7 digitos (lo asigna AIR-E).")
        if self.buscar(suscriptor):
            raise ValueError("Ese suscriptor ya existe.")
        donante = self.tarifario().get(uso)
        if donante is None:
            raise ValueError(f"No hay donante para el uso '{uso}'.")
        centavos, don = donante
        xml = don.xml
        # historico en cero para usuario nuevo
        for i in range(1, 7):
            xml = poner_tag(xml, f"MES_{i}", "0,00")
        r = Registro(suscriptor, centavos, don.periodo, xml, lote)
        self.registros[(lote, suscriptor)] = r
        self.cambios.append(f"ALTA {suscriptor} ({lote}) como {uso}")

    def baja(self, suscriptor):
        regs = self.buscar(suscriptor)
        if not regs:
            raise ValueError("Suscriptor no encontrado.")
        for r in regs:
            del self.registros[(r.lote, r.suscriptor)]
            self.cambios.append(f"BAJA {suscriptor} ({r.lote})")

    # ---------------- periodo siguiente ----------------
    def avanzar_periodo(self):
        """Prepara el mes siguiente: corre el historico (MES_1 toma el total
        actual), actualiza PERIODO_FACTURADO y el periodo de las lineas FC."""
        for r in self.registros.values():
            historico = [tag(r.xml, f"MES_{i}") for i in range(1, 7)]
            nuevos = [a_texto(r.total)] + historico[:5]
            for i, v in enumerate(nuevos, start=1):
                r.xml = poner_tag(r.xml, f"MES_{i}", v)
            anio, mes = int(r.periodo[:4]), int(r.periodo[4:])
            mes += 1
            if mes > 12:
                mes, anio = 1, anio + 1
            r.periodo = f"{anio}{mes:02d}"
            r.xml = poner_tag(r.xml, "PERIODO_FACTURADO", f"{MESES[mes-1]} {anio}")
        anio, mes = int(self.periodo_archivo[:4]), int(self.periodo_archivo[4:])
        mes += 1
        if mes > 12:
            mes, anio = 1, anio + 1
        self.periodo_archivo = f"{anio}{mes:02d}"
        self.cambios.append(f"Periodo avanzado a {self.periodo_archivo} (archivos) ")

    # ---------------- validacion y exportacion ----------------
    def validar(self):
        errores = []
        tarifas = {u: c for u, (c, _) in self.tarifario().items()}
        for r in self.registros.values():
            try:
                if len(r.fc()) != LARGO_FC:
                    errores.append(f"{r.suscriptor}: FC de largo invalido")
                if len(r.fd()) != LARGO_FD:
                    errores.append(f"{r.suscriptor}: FD de largo invalido")
            except ValueError as e:
                errores.append(str(e))
            if int(round(r.total * 100)) != r.centavos:
                errores.append(f"{r.suscriptor}: TOTAL del XML ({r.total}) no cuadra con el valor FC ({r.centavos/100})")
            if tarifas.get(r.uso) != r.centavos:
                errores.append(f"{r.suscriptor}: valor distinto a la tarifa de '{r.uso}'")
        return errores

    def resumen(self):
        print("\n----- RESUMEN -----")
        for lote in ("principal", "hoja1"):
            regs = [r for r in self.registros.values() if r.lote == lote]
            if not regs:
                continue
            total = sum(r.centavos for r in regs) / 100
            print(f"Lote {lote}: {len(regs)} facturas | Total $ {a_texto(total)}")
            for uso, n in sorted(Counter(r.uso for r in regs).items()):
                print(f"   {uso:<28} {n:>5}")
        if self.cambios:
            print("Cambios pendientes de exportar:")
            for c in self.cambios:
                print("   *", c)

    def exportar(self, carpeta_salida):
        os.makedirs(carpeta_salida, exist_ok=True)
        p = self.periodo_archivo
        nombres = {
            "principal": (f"INGESAM_VOLCADO_{CONVENIO}_{p}.txt",
                          f"INFO_ADICIONAL_INGESAM_{CONVENIO}_{p}.txt"),
            "hoja1": (f"Hoja1_INGESAM_VOLCADO_{CONVENIO}_{p}.txt",
                      f"Hoja1_INFO_ADICIONAL_INGESAM_{CONVENIO}_{p}.txt"),
        }
        for lote, (f_vol, f_info) in nombres.items():
            regs = [r for r in self.registros.values() if r.lote == lote]
            if not regs:
                continue
            with open(os.path.join(carpeta_salida, f_vol), "w",
                      encoding="ascii", newline="") as fv, \
                 open(os.path.join(carpeta_salida, f_info), "w",
                      encoding="ascii", newline="") as fi:
                for r in regs:
                    fv.write(r.fc() + "\n")
                    fi.write(r.fd() + "\n")
            print(f"  Exportado {f_vol} y {f_info} ({len(regs)} facturas)")
        self._exportar_xlsx(carpeta_salida)
        self.cambios.clear()

    def _exportar_xlsx(self, carpeta_salida):
        """Actualiza la columna PRECIO de TABLA_PRECIOS_ASEO.xlsx con las
        tarifas vigentes en los registros (lo facturado a AIR-E)."""
        origen = os.path.join(self.carpeta, "TABLA_PRECIOS_ASEO.xlsx")
        destino = os.path.join(carpeta_salida, "TABLA_PRECIOS_ASEO.xlsx")
        if not os.path.exists(origen):
            print("  AVISO: no se encontro TABLA_PRECIOS_ASEO.xlsx, no se exporta.")
            return
        try:
            from openpyxl import load_workbook
        except ImportError:
            shutil.copy(origen, destino)
            print("  AVISO: openpyxl no instalado; xlsx copiado sin cambios.")
            return
        wb = load_workbook(origen)
        ws = wb.active
        cols = {str(c.value).strip(): c.column for c in ws[1] if c.value}
        col_sub = cols.get("SUBCATEGORIA")
        col_precio = next((v for k, v in cols.items() if k.startswith("PRECIO")), None)
        tarifas = {u: c / 100 for u, (c, _) in self.tarifario().items()}
        for fila in range(2, ws.max_row + 1):
            sub = ws.cell(row=fila, column=col_sub).value
            if sub and str(sub).strip() in tarifas:
                ws.cell(row=fila, column=col_precio).value = tarifas[str(sub).strip()]
        wb.save(destino)
        print(f"  Exportado TABLA_PRECIOS_ASEO.xlsx")


# ----------------------------------------------------------------------
# interfaz de consola
# ----------------------------------------------------------------------
def elegir_uso(v):
    usos = list(v.tarifario().keys())
    for i, u in enumerate(usos, 1):
        cent, _ = v.tarifario()[u]
        print(f"  {i}. {u:<28} $ {a_texto(cent/100)}")
    n = input("Numero de subcategoria: ").strip()
    return usos[int(n) - 1]


def main():
    carpeta = sys.argv[1] if len(sys.argv) > 1 else "."
    archivos = [f for f in os.listdir(carpeta)
                if re.match(rf"INGESAM_VOLCADO_{CONVENIO}_\d{{6}}\.txt$", f)]
    if not archivos:
        print("No se encontro ningun INGESAM_VOLCADO en la carpeta. Usa: python editor_volcado.py <carpeta>")
        return
    periodo = sorted(re.search(r"(\d{6})", f).group(1) for f in archivos)[-1]
    print(f"Cargando periodo {periodo} desde '{carpeta}'...")
    v = Volcado(carpeta, periodo)
    v.resumen()

    menu = """
================ EDITOR DE VOLCADO AIR-E ================
 1. Buscar usuario
 2. Cambiar subcategoria (estrato) de un usuario
 3. Cambiar tarifa de una subcategoria
 4. Alta de usuario nuevo
 5. Baja de usuario
 6. Avanzar al periodo siguiente (corre historico)
 7. Validar
 8. Resumen
 9. Exportar archivos
 0. Salir
=========================================================="""
    while True:
        print(menu)
        op = input("Opcion: ").strip()
        try:
            if op == "1":
                s = input("Suscriptor: ").strip()
                for r in v.buscar(s) or [print("No encontrado.")] and []:
                    print(f"  [{r.lote}] {r.suscriptor} | {r.uso} | $ {a_texto(r.centavos/100)} | periodo {r.periodo}")
            elif op == "2":
                s = input("Suscriptor: ").strip()
                regs = v.buscar(s)
                if not regs:
                    print("No encontrado."); continue
                print(f"Subcategoria actual: {regs[0].uso}. Nueva:")
                v.cambiar_uso(s, elegir_uso(v))
                print("Listo.")
            elif op == "3":
                print("Subcategoria a modificar:")
                uso = elegir_uso(v)
                t = float(input(f"Nuevo total para '{uso}' (ej. 8860): ").replace(",", "."))
                v.cambiar_tarifa(uso, t)
                print("Listo.")
            elif op == "4":
                s = input("Codigo del suscriptor nuevo (solo digitos): ").strip()
                lote = "hoja1" if input("Lote [1=principal, 2=hoja1]: ").strip() == "2" else "principal"
                print("Subcategoria:")
                v.alta(s, elegir_uso(v), lote)
                print("Listo.")
            elif op == "5":
                v.baja(input("Suscriptor a retirar: ").strip())
                print("Listo.")
            elif op == "6":
                if input("Esto corre el historico y avanza el mes. Confirmar (s/n): ").lower() == "s":
                    v.avanzar_periodo()
                    print(f"Periodo nuevo: {v.periodo_archivo}")
            elif op == "7":
                e = v.validar()
                print("Sin errores. Todo cuadra." if not e else f"{len(e)} ERRORES:")
                for x in e[:20]:
                    print("  -", x)
            elif op == "8":
                v.resumen()
            elif op == "9":
                e = v.validar()
                if e:
                    print(f"OJO: hay {len(e)} errores de validacion. Corrigelos antes de enviar.")
                    if input("Exportar de todas formas (s/n): ").lower() != "s":
                        continue
                destino = input("Carpeta de salida [salida]: ").strip() or "salida"
                v.exportar(destino)
                print(f"Archivos listos en '{destino}' para enviar a AIR-E.")
            elif op == "0":
                break
        except (ValueError, IndexError) as e:
            print("ERROR:", e)


if __name__ == "__main__":
    main()
