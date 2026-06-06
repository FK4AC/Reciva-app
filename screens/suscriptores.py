import os
from datetime import date

from kivy.uix.screenmanager import Screen
from kivy.uix.popup import Popup
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.spinner import Spinner
from kivy.uix.scrollview import ScrollView
from kivy.graphics import Color, Rectangle, RoundedRectangle
from kivy.properties import StringProperty
from kivy.clock import Clock
import threading
from db.connection import get_connection
import utils.overlay as overlay
from utils.estado_cuenta import generar_estado_cuenta
from theme import (TINTA, BG, STAGE, CARD, VERMILLON, LADRILLO,
                   LINE, MUTED, TEXT_SEC, SUCCESS, WARNING, DANGER, SIDEBAR_BTN)

MESES = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
    5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
    9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
}


class SuscriptoresScreen(Screen):
    mensaje = StringProperty('')

    def on_enter(self):
        self.ids.lista_suscriptores.clear_widgets()
        self.mensaje = 'Usa el buscador o presiona "Mostrar todos"'

    def buscar(self, texto):
        filtro = texto.strip()
        self.ids.lista_suscriptores.clear_widgets()
        self.mensaje = ''
        overlay.show('Buscando…')
        threading.Thread(target=lambda: self._tarea_lista(filtro), daemon=True).start()

    def mostrar_todos(self):
        self.ids.lista_suscriptores.clear_widgets()
        self.mensaje = ''
        overlay.show('Cargando…')
        threading.Thread(target=lambda: self._tarea_lista(''), daemon=True).start()

    def _tarea_lista(self, filtro):
        rows, error = [], None
        conn = get_connection()
        if not conn:
            Clock.schedule_once(lambda *_: self._aplicar_lista([], 'Error de conexión'), 0)
            return
        cursor = conn.cursor()
        try:
            if filtro:
                like = f'%{filtro}%'
                cursor.execute("""
                    SELECT cuenta, nombre, barrio, estrato, estado_suministro
                    FROM suscriptores
                    WHERE nombre LIKE %s OR CAST(cuenta AS CHAR) LIKE %s
                    ORDER BY nombre LIMIT 300
                """, (like, like))
            else:
                cursor.execute("""
                    SELECT cuenta, nombre, barrio, estrato, estado_suministro
                    FROM suscriptores ORDER BY nombre LIMIT 300
                """)
            rows = cursor.fetchall()
        except Exception as e:
            error = str(e)
        finally:
            cursor.close()
            conn.close()
        Clock.schedule_once(lambda *_: self._aplicar_lista(rows, error), 0)

    def _aplicar_lista(self, rows, error):
        overlay.hide()
        if error:
            if 'conexión' in error.lower() or error == 'Error de conexión':
                from kivy.app import App
                App.get_running_app().ir_sin_conexion(self.name)
            else:
                self.mensaje = f'Error: {error}'
            return
        lista = self.ids.lista_suscriptores
        total = len(rows)
        self.mensaje = f'{total} suscriptores encontrados' + (' (máx. 300)' if total == 300 else '')
        for i, (cuenta, nombre, barrio, estrato, estado) in enumerate(rows):
            lista.add_widget(self._fila(cuenta, nombre, barrio, estrato, estado, i))

    def _fila(self, cuenta, nombre, barrio, estrato, estado, idx=0):
        fila = BoxLayout(orientation='horizontal', size_hint_y=None, height=38, spacing=2)
        bg = CARD if idx % 2 == 0 else STAGE

        with fila.canvas.before:
            Color(*bg)
            rect = Rectangle(pos=fila.pos, size=fila.size)

        fila.bind(pos=lambda _, v, r=rect: setattr(r, 'pos', v))
        fila.bind(size=lambda _, v, r=rect: setattr(r, 'size', v))

        datos = [
            (str(cuenta),           0.18),
            ((nombre or '')[:35],   0.36),
            ((barrio or '-')[:20],  0.18),
            (str(estrato or '-'),   0.09),
            (str(estado or '-'),    0.12),
        ]
        for texto, sx in datos:
            lbl = Label(text=texto, size_hint_x=sx, font_size=13,
                        color=TINTA, halign='left', valign='middle')
            lbl.bind(size=lambda inst, v: setattr(inst, 'text_size', (v[0] - 4, v[1])))
            fila.add_widget(lbl)

        btn = Button(text='Ver', size_hint_x=0.07, font_size=12,
                     background_normal='', background_color=VERMILLON, color=(1,1,1,1))
        btn.bind(on_press=lambda x, c=cuenta: self._ver_detalle(c))
        fila.add_widget(btn)

        return fila

    def _ver_detalle(self, cuenta):
        overlay.show('Cargando suscriptor…')

        def _tarea():
            info, facturas, recaudos, recaudos_det, pqr_list, error = \
                None, {}, {}, {}, [], None
            conn = get_connection()
            if not conn:
                Clock.schedule_once(
                    lambda *_: (overlay.hide(),
                                setattr(self, 'mensaje', 'Error de conexión')), 0)
                return
            cursor = conn.cursor()
            try:
                cursor.execute("""
                    SELECT cuenta, nombre, direccion, barrio, estrato,
                           estado_suministro, municipio, susccodi
                    FROM suscriptores WHERE cuenta = %s LIMIT 1
                """, (cuenta,))
                info = cursor.fetchone()
                cursor.execute("""
                    SELECT año, mes, SUM(valor_recibo)
                    FROM facturas WHERE cuenta_contrato = %s
                    GROUP BY año, mes ORDER BY año, mes
                """, (cuenta,))
                facturas = {(int(r[0]), int(r[1])): float(r[2]) for r in cursor.fetchall()}
                # Recaudos individuales para el desglose por factura
                cursor.execute("""
                    SELECT año, mes, numero_factura, fecha_recaudo,
                           valor_recibo, concepto
                    FROM recaudos WHERE cuenta_contrato = %s
                    ORDER BY año, mes, fecha_recaudo
                """, (cuenta,))
                for r in cursor.fetchall():
                    key = (int(r[0]), int(r[1]))
                    recaudos[key] = recaudos.get(key, 0) + float(r[4])
                    recaudos_det.setdefault(key, []).append(
                        (r[2], r[3], float(r[4]), str(r[5] or ''))
                    )
                cursor.execute("""
                    SELECT id, tipo, asunto, estado, fecha_creacion
                    FROM pqr WHERE cuenta_contrato = %s
                    ORDER BY fecha_creacion DESC
                """, (cuenta,))
                pqr_list = cursor.fetchall()
            except Exception as e:
                error = str(e)
            finally:
                cursor.close()
                conn.close()
            Clock.schedule_once(
                lambda *_: _aplicar(info, facturas, recaudos, recaudos_det,
                                    pqr_list, error), 0)

        def _aplicar(info, facturas, recaudos, recaudos_det, pqr_list, error):
            overlay.hide()
            if error:
                self.mensaje = f'Error al cargar detalle: {error}'
                return
            self._popup_detalle(info, facturas, recaudos, recaudos_det, pqr_list)

        threading.Thread(target=_tarea, daemon=True).start()

    def _popup_detalle(self, info, facturas, recaudos, recaudos_det, pqr_list):
        if not info:
            return

        cuenta, nombre, direccion, barrio, estrato, estado_sum, municipio, susccodi = info

        total_facturado = sum(facturas.values())
        total_pagado    = sum(recaudos.values())
        deuda_total     = max(0, total_facturado - total_pagado)

        meses_sin_pagar = [
            (a, m) for (a, m), v in sorted(facturas.items())
            if recaudos.get((a, m), 0) < v * 0.95
        ]
        consecutivos = 0
        for (a, m), v in sorted(facturas.items(), reverse=True):
            if recaudos.get((a, m), 0) < v * 0.95:
                consecutivos += 1
            else:
                break

        n_sin = len(meses_sin_pagar)
        if n_sin == 0:
            color_estado = SUCCESS
            texto_estado = 'Al día'
        elif n_sin <= 2:
            color_estado = WARNING
            texto_estado = f'{n_sin} mes{"" if n_sin == 1 else "es"} sin pagar'
        else:
            color_estado = DANGER
            texto_estado = f'{n_sin} meses sin pagar'

        # ── Root ──
        content = BoxLayout(orientation='vertical', spacing=0)
        with content.canvas.before:
            Color(*CARD)
            _bg = Rectangle(pos=content.pos, size=content.size)
        content.bind(pos=lambda _, v: setattr(_bg, 'pos', v),
                     size=lambda _, v: setattr(_bg, 'size', v))

        # ── Franja superior TINTA ──
        top = BoxLayout(orientation='vertical', size_hint_y=None, height=80,
                        padding=[18, 8])
        with top.canvas.before:
            Color(*TINTA)
            _top = Rectangle(pos=top.pos, size=top.size)
        top.bind(pos=lambda _, v: setattr(_top, 'pos', v),
                 size=lambda _, v: setattr(_top, 'size', v))

        name_row = BoxLayout(size_hint_y=None, height=36)
        name_row.add_widget(Label(
            text=nombre or '—', bold=True, font_size=17,
            color=(1, 1, 1, 1), halign='left', valign='middle', size_hint_x=0.62,
        ))
        chip_wrap = BoxLayout(size_hint_x=0.38, padding=[10, 4, 0, 4])
        chip = BoxLayout()
        with chip.canvas.before:
            Color(*color_estado)
            _chip = RoundedRectangle(pos=chip.pos, size=chip.size, radius=[13])
        chip.bind(pos=lambda _, v, r=_chip: setattr(r, 'pos', v),
                  size=lambda _, v, r=_chip: setattr(r, 'size', v))
        chip.add_widget(Label(text=texto_estado, color=(1, 1, 1, 1), bold=True, font_size=11))
        chip_wrap.add_widget(chip)
        name_row.add_widget(chip_wrap)
        top.add_widget(name_row)

        meta_parts = [f'#{cuenta}', barrio or '—',
                      f'Estrato {estrato or "—"}', estado_sum or '—']
        top.add_widget(Label(
            text='  ·  '.join(meta_parts), font_size=11,
            color=LINE, halign='left', valign='middle',
            size_hint_y=None, height=26,
        ))
        content.add_widget(top)

        # ── 3 fichas financieras ──
        fichas = BoxLayout(size_hint_y=None, height=68, spacing=1)
        for lbl_t, val_t, col in [
            ('Total Facturado', f'${total_facturado:,.0f}', TEXT_SEC),
            ('Total Pagado',    f'${total_pagado:,.0f}',    SUCCESS),
            ('Deuda Total',     f'${deuda_total:,.0f}',     DANGER if deuda_total > 0 else MUTED),
        ]:
            f = BoxLayout(orientation='vertical', padding=[0, 10])
            with f.canvas.before:
                Color(*STAGE)
                _fr = Rectangle(pos=f.pos, size=f.size)
            f.bind(pos=lambda _, v, r=_fr: setattr(r, 'pos', v),
                   size=lambda _, v, r=_fr: setattr(r, 'size', v))
            f.add_widget(Label(text=val_t, font_size=17, bold=True, color=col))
            f.add_widget(Label(text=lbl_t, font_size=10, color=MUTED))
            fichas.add_widget(f)
        content.add_widget(fichas)

        # ── Encabezado tabla ──
        hdr = BoxLayout(size_hint_y=None, height=30, spacing=2)
        with hdr.canvas.before:
            Color(*TINTA)
            _hdr = Rectangle(pos=hdr.pos, size=hdr.size)
        hdr.bind(pos=lambda _, v: setattr(_hdr, 'pos', v),
                 size=lambda _, v: setattr(_hdr, 'size', v))
        for txt, sx, al in [
            ('Año',       0.10, 'left'),
            ('Mes',       0.18, 'left'),
            ('Facturado', 0.22, 'right'),
            ('Pagado',    0.22, 'right'),
            ('Estado',    0.18, 'center'),
            ('',          0.10, 'left'),
        ]:
            lbl_h = Label(text=txt, bold=True, size_hint_x=sx,
                          font_size=10, color=LINE, halign=al, valign='middle')
            lbl_h.bind(size=lambda inst, v: setattr(inst, 'text_size', (v[0]-6, v[1])))
            hdr.add_widget(lbl_h)
        content.add_widget(hdr)

        # ── Tabla scrollable con filas expandibles ──
        scroll = ScrollView()
        tabla  = BoxLayout(orientation='vertical', size_hint_y=None, spacing=0)
        tabla.bind(minimum_height=tabla.setter('height'))

        ROW_H   = 36   # altura fila principal
        SUB_HDR = 24   # encabezado sub-panel
        SUB_ROW = 26   # fila de recaudo individual

        all_months = sorted(set(list(facturas.keys()) + list(recaudos.keys())))
        for i, (a, m) in enumerate(all_months):
            fac_val  = facturas.get((a, m), 0)
            rec_val  = recaudos.get((a, m), 0)
            det_rows = recaudos_det.get((a, m), [])
            row_bg   = CARD if i % 2 == 0 else STAGE

            if fac_val > 0 and rec_val >= fac_val * 0.95:
                est_txt, est_col = 'Pagado', SUCCESS
            elif fac_val > 0:
                est_txt, est_col = 'Pendiente', DANGER
            else:
                est_txt, est_col = 'Sin factura', MUTED

            tiene_det = bool(det_rows)
            sub_h     = SUB_HDR + len(det_rows) * SUB_ROW if tiene_det else 0

            # ── Grupo: fila principal + divisor + sub-panel ──
            grupo = BoxLayout(orientation='vertical', size_hint_y=None,
                              height=ROW_H)

            # Fila principal
            fila = BoxLayout(size_hint_y=None, height=ROW_H, padding=[0, 0, 0, 0])
            with fila.canvas.before:
                Color(*row_bg)
                _rf = Rectangle(pos=fila.pos, size=fila.size)
            fila.bind(pos=lambda _, v, r=_rf: setattr(r, 'pos', v),
                      size=lambda _, v, r=_rf: setattr(r, 'size', v))

            for val, sx, col, al in [
                (str(a),                    0.10, TEXT_SEC,                           'left'),
                (MESES.get(m, str(m)),      0.18, TINTA,                              'left'),
                (f'${fac_val:,.0f}',        0.22, TEXT_SEC,                           'right'),
                (f'${rec_val:,.0f}',        0.22, SUCCESS if rec_val > 0 else MUTED,  'right'),
                (est_txt,                   0.18, est_col,                             'center'),
            ]:
                lbl = Label(text=val, color=col, font_size=12, size_hint_x=sx,
                            halign=al, valign='middle')
                lbl.bind(size=lambda inst, v: setattr(inst, 'text_size', (v[0]-6, v[1])))
                fila.add_widget(lbl)

            # Botón expandir / colapsar (sólo si hay desglose)
            btn_exp = Button(
                text='▼' if tiene_det else '',
                size_hint_x=0.10, font_size=10,
                background_normal='', background_color=(0, 0, 0, 0),
                color=MUTED, disabled=not tiene_det,
            )
            fila.add_widget(btn_exp)
            grupo.add_widget(fila)

            # Sub-panel (construido siempre, oculto con height=0 + opacity=0)
            sub = BoxLayout(orientation='vertical', size_hint_y=None,
                            height=0, opacity=0)

            if tiene_det:
                # Encabezado del sub-panel
                sh = BoxLayout(size_hint_y=None, height=SUB_HDR, padding=[32, 0, 8, 0])
                with sh.canvas.before:
                    Color(*STAGE)
                    _sh = Rectangle(pos=sh.pos, size=sh.size)
                sh.bind(pos=lambda _, v, r=_sh: setattr(r, 'pos', v),
                        size=lambda _, v, r=_sh: setattr(r, 'size', v))
                for txt, sx in [('N° Factura', 0.28), ('Fecha pago', 0.22),
                                 ('Concepto',   0.32), ('Valor',      0.18)]:
                    sh.add_widget(Label(
                        text=txt, font_size=9, bold=True, color=MUTED,
                        size_hint_x=sx, halign='left', valign='middle',
                    ))
                sub.add_widget(sh)

                for j, (nf, fecha_rec, val_rec, concepto) in enumerate(det_rows):
                    sf_bg = CARD if j % 2 == 0 else (0.94, 0.91, 0.87, 1)
                    sf = BoxLayout(size_hint_y=None, height=SUB_ROW, padding=[32, 0, 8, 0])
                    with sf.canvas.before:
                        Color(*sf_bg)
                        _sf = Rectangle(pos=sf.pos, size=sf.size)
                    sf.bind(pos=lambda _, v, r=_sf: setattr(r, 'pos', v),
                            size=lambda _, v, r=_sf: setattr(r, 'size', v))
                    nf_str    = str(nf) if nf else '—'
                    fecha_str = str(fecha_rec)[:10] if fecha_rec else '—'
                    conc_str  = (concepto[:24] + '…') if len(concepto) > 24 else concepto
                    for txt, sx, col in [
                        (nf_str,              0.28, TEXT_SEC),
                        (fecha_str,           0.22, TINTA),
                        (conc_str,            0.32, MUTED),
                        (f'${val_rec:,.0f}',  0.18, SUCCESS),
                    ]:
                        lbl = Label(text=txt, font_size=10, color=col,
                                    size_hint_x=sx, halign='left', valign='middle')
                        lbl.bind(size=lambda inst, v: setattr(inst, 'text_size',
                                                               (v[0]-4, v[1])))
                        sf.add_widget(lbl)
                    sub.add_widget(sf)

            grupo.add_widget(sub)

            # Divisor
            div = BoxLayout(size_hint_y=None, height=1)
            with div.canvas.before:
                Color(*LINE)
                Rectangle(pos=div.pos, size=div.size)
            grupo.add_widget(div)

            def _toggle(_, sub=sub, grupo=grupo, sub_h=sub_h,
                        btn=btn_exp, row_h=ROW_H):
                if sub.height == 0:
                    sub.height  = sub_h
                    sub.opacity = 1
                    grupo.height = row_h + sub_h + 1
                    btn.text = '▲'
                else:
                    sub.height  = 0
                    sub.opacity = 0
                    grupo.height = row_h + 1
                    btn.text = '▼'

            btn_exp.bind(on_press=_toggle)
            tabla.add_widget(grupo)

        scroll.add_widget(tabla)
        content.add_widget(scroll)

        # ── Pie ──
        pie = BoxLayout(size_hint_y=None, height=52, spacing=10, padding=[14, 8])
        with pie.canvas.before:
            Color(*STAGE)
            _pie = Rectangle(pos=pie.pos, size=pie.size)
        pie.bind(pos=lambda _, v: setattr(_pie, 'pos', v),
                 size=lambda _, v: setattr(_pie, 'size', v))

        pie.add_widget(Label(
            text=f'{consecutivos} mes(es) consecutivo(s) sin pagar',
            font_size=11, color=MUTED, halign='left',
        ))
        lbl_pdf = Label(text='', font_size=11, color=SUCCESS,
                        size_hint_x=None, width=180)
        pie.add_widget(lbl_pdf)

        def _pill(text, bg, fg=(1, 1, 1, 1), w=115):
            b = Button(
                text=text, size_hint_x=None, width=w, font_size=12,
                color=fg, background_color=(0, 0, 0, 0),
                background_normal='', background_down='',
            )
            with b.canvas.before:
                Color(*bg)
                rr = RoundedRectangle(pos=b.pos, size=b.size, radius=[20])
            b.bind(pos=lambda _, v, r=rr: setattr(r, 'pos', v),
                   size=lambda _, v, r=rr: setattr(r, 'size', v))
            return b

        btn_editar = _pill('Editar',      WARNING)
        btn_pdf    = _pill('Generar PDF', TINTA, w=130)
        btn_cerrar = _pill('Cerrar',      LINE,  fg=TINTA)

        pie.add_widget(btn_editar)
        pie.add_widget(btn_pdf)
        pie.add_widget(btn_cerrar)
        content.add_widget(pie)

        popup = Popup(
            title=f'Suscriptor — {nombre}',
            content=content,
            size_hint=(0.82, 0.88),
            background_color=VERMILLON,
            title_color=(1, 1, 1, 1),
            separator_color=LINE,
        )

        def _abrir_selector_rango(_):
            periodos_todos = sorted(set(list(facturas.keys()) + list(recaudos.keys())))
            if not periodos_todos:
                try:
                    ruta = generar_estado_cuenta(info, facturas, recaudos, pqr_list)
                    lbl_pdf.text = 'PDF guardado'
                    os.startfile(ruta)
                except Exception as exc:
                    lbl_pdf.text = f'Error: {exc}'
                    lbl_pdf.color = DANGER
                return

            def p_str(a, m):
                return f"{MESES.get(m, str(m))} {a}"

            str_a_key = {p_str(a, m): (a, m) for a, m in periodos_todos}
            opciones  = [p_str(a, m) for a, m in periodos_todos]

            # ── Popup de rango ────────────────────────────────────────
            rng = BoxLayout(orientation='vertical', spacing=0)
            with rng.canvas.before:
                Color(*CARD)
                r_bg = Rectangle(pos=rng.pos, size=rng.size)
            rng.bind(pos=lambda _, v: setattr(r_bg, 'pos', v),
                     size=lambda _, v: setattr(r_bg, 'size', v))

            rng_hdr = BoxLayout(size_hint_y=None, height=52, padding=[16, 0])
            with rng_hdr.canvas.before:
                Color(*TINTA)
                r_h = Rectangle(pos=rng_hdr.pos, size=rng_hdr.size)
            rng_hdr.bind(pos=lambda _, v: setattr(r_h, 'pos', v),
                         size=lambda _, v: setattr(r_h, 'size', v))
            rng_hdr.add_widget(Label(
                text='Rango del Estado de Cuenta', bold=True, font_size=14,
                color=(1, 1, 1, 1), halign='left', valign='middle',
                text_size=(400, 52)
            ))
            rng.add_widget(rng_hdr)

            body = BoxLayout(orientation='vertical', padding=[20, 16], spacing=12)
            rng.add_widget(body)
            body.add_widget(Label(
                text='El PDF incluirá únicamente los períodos dentro del rango.',
                font_size=11, color=MUTED,
                size_hint_y=None, height=20,
                halign='left', text_size=(600, 20)
            ))

            def _fila_sp(lbl_txt, default):
                fila = BoxLayout(size_hint_y=None, height=40, spacing=10)
                fila.add_widget(Label(
                    text=lbl_txt, size_hint_x=None, width=54,
                    font_size=12, color=MUTED,
                    halign='right', valign='middle', text_size=(54, 40)
                ))
                wrap = BoxLayout()
                with wrap.canvas.before:
                    Color(*STAGE)
                    rr = RoundedRectangle(pos=wrap.pos, size=wrap.size, radius=[14])
                wrap.bind(pos=lambda _, v, r=rr: setattr(r, 'pos', v),
                          size=lambda _, v, r=rr: setattr(r, 'size', v))
                sp = Spinner(text=default, values=opciones,
                             background_color=(0, 0, 0, 0),
                             background_normal='', color=TINTA, font_size=12)
                wrap.add_widget(sp)
                fila.add_widget(wrap)
                return fila, sp

            fila_d, sp_desde = _fila_sp('Desde:', opciones[0])
            fila_h, sp_hasta = _fila_sp('Hasta:', opciones[-1])
            body.add_widget(fila_d)
            body.add_widget(fila_h)

            pie_rng = BoxLayout(size_hint_y=None, height=52, spacing=10, padding=[16, 8])
            with pie_rng.canvas.before:
                Color(*STAGE)
                r_ft = Rectangle(pos=pie_rng.pos, size=pie_rng.size)
            pie_rng.bind(pos=lambda _, v: setattr(r_ft, 'pos', v),
                         size=lambda _, v: setattr(r_ft, 'size', v))
            rng.add_widget(pie_rng)

            rng_popup = Popup(title='', content=rng,
                              size_hint=(0.38, 0.44),
                              background_color=CARD, separator_height=0)

            def _generar(_):
                rng_popup.dismiss()
                desde_key = str_a_key.get(sp_desde.text)
                hasta_key  = str_a_key.get(sp_hasta.text)
                if desde_key and hasta_key:
                    if desde_key > hasta_key:
                        desde_key, hasta_key = hasta_key, desde_key
                    fac_f = {k: v for k, v in facturas.items() if desde_key <= k <= hasta_key}
                    rec_f = {k: v for k, v in recaudos.items() if desde_key <= k <= hasta_key}
                else:
                    fac_f, rec_f = facturas, recaudos
                try:
                    ruta = generar_estado_cuenta(info, fac_f, rec_f, pqr_list)
                    lbl_pdf.text = f'{sp_desde.text} → {sp_hasta.text}'
                    os.startfile(ruta)
                except Exception as exc:
                    lbl_pdf.text = f'Error: {exc}'
                    lbl_pdf.color = DANGER

            btn_cancelar_rng = _pill('Cancelar',    LINE,  fg=TINTA, w=110)
            btn_generar_rng  = _pill('Generar PDF', TINTA, w=130)
            btn_cancelar_rng.bind(on_press=rng_popup.dismiss)
            btn_generar_rng.bind(on_press=_generar)
            pie_rng.add_widget(btn_cancelar_rng)
            pie_rng.add_widget(btn_generar_rng)
            rng_popup.open()

        def abrir_editor(_):
            popup.dismiss()
            self._popup_formulario(cuenta_editar=cuenta)

        btn_editar.bind(on_press=abrir_editor)
        btn_pdf.bind(on_press=_abrir_selector_rango)
        btn_cerrar.bind(on_press=popup.dismiss)
        popup.open()

    # ------------------------------------------------------------------
    #  Popup compartido: Nuevo / Editar suscriptor
    # ------------------------------------------------------------------
    def nuevo_suscriptor(self):
        self._popup_formulario(cuenta_editar=None)

    def _popup_formulario(self, cuenta_editar=None):
        es_edicion = cuenta_editar is not None
        datos_actuales = {}

        if es_edicion:
            conn = get_connection()
            if not conn:
                return
            cur = conn.cursor()
            try:
                cur.execute("""
                    SELECT cuenta, susccodi, nombre, direccion, municipio,
                           barrio, subcategoria, estrato, estado_suministro
                    FROM suscriptores WHERE cuenta = %s LIMIT 1
                """, (cuenta_editar,))
                row = cur.fetchone()
                if row:
                    keys = ['cuenta', 'susccodi', 'nombre', 'direccion',
                            'municipio', 'barrio', 'subcategoria',
                            'estrato', 'estado_suministro']
                    datos_actuales = {k: (str(v) if v is not None else '')
                                      for k, v in zip(keys, row)}
            finally:
                cur.close()
                conn.close()

        content = BoxLayout(orientation='vertical', spacing=10, padding=15)
        with content.canvas.before:
            Color(*CARD)
            _cbg = Rectangle(pos=content.pos, size=content.size)
        content.bind(pos=lambda _, v: setattr(_cbg, 'pos', v),
                     size=lambda _, v: setattr(_cbg, 'size', v))

        form = GridLayout(cols=2, size_hint_y=None, spacing=8, padding=[0, 4])
        form.bind(minimum_height=form.setter('height'))

        campos = [
            ('Cuenta *',          'cuenta',           not es_edicion),
            ('SUSCCODI',          'susccodi',          not es_edicion),
            ('Nombre *',          'nombre',            True),
            ('Dirección',         'direccion',         True),
            ('Municipio',         'municipio',         True),
            ('Barrio',            'barrio',            True),
            ('Subcategoría',      'subcategoria',      True),
            ('Estrato',           'estrato',           True),
            ('Estado suministro', 'estado_suministro', True),
        ]

        inputs = {}
        for label_txt, key, editable in campos:
            lbl = Label(
                text=label_txt, color=VERMILLON,
                bold=True, halign='right', size_hint_y=None, height=38,
            )
            lbl.bind(size=lambda inst, v: setattr(inst, 'text_size', (v[0], v[1])))
            form.add_widget(lbl)

            inp = TextInput(
                text=datos_actuales.get(key, ''),
                multiline=False,
                size_hint_y=None, height=38,
                readonly=not editable,
                background_color=STAGE if not editable else CARD,
                foreground_color=MUTED  if not editable else TINTA,
                cursor_color=VERMILLON,
                padding=[8, 10],
            )
            form.add_widget(inp)
            inputs[key] = inp

        scroll = ScrollView(size_hint_y=1)
        scroll.add_widget(form)
        content.add_widget(scroll)

        lbl_err = Label(text='', color=DANGER, size_hint_y=None, height=26)
        content.add_widget(lbl_err)

        btns = BoxLayout(size_hint_y=None, height=44, spacing=12)
        lbl_ok = Label(text='', color=SUCCESS, font_size=12)

        def _pill_f(text, bg, fg=(1, 1, 1, 1)):
            b = Button(
                text=text, font_size=12, color=fg,
                background_color=(0, 0, 0, 0),
                background_normal='', background_down='',
            )
            with b.canvas.before:
                Color(*bg)
                rr = RoundedRectangle(pos=b.pos, size=b.size, radius=[20])
            b.bind(pos=lambda _, v, r=rr: setattr(r, 'pos', v),
                   size=lambda _, v, r=rr: setattr(r, 'size', v))
            return b

        btn_guardar  = _pill_f('Guardar',  SUCCESS)
        btn_cancelar = _pill_f('Cancelar', LINE, fg=TINTA)

        btns.add_widget(lbl_ok)
        btns.add_widget(btn_guardar)
        btns.add_widget(btn_cancelar)
        content.add_widget(btns)

        titulo = f'Editar Suscriptor — {cuenta_editar}' if es_edicion else 'Nuevo Suscriptor'
        popup = Popup(
            title=titulo, content=content, size_hint=(0.55, 0.82),
            background_color=VERMILLON, title_color=(1, 1, 1, 1),
            separator_color=LINE,
        )

        def guardar(_):
            nombre_val = inputs['nombre'].text.strip()
            if not nombre_val:
                lbl_err.text = 'El nombre es obligatorio'
                return

            conn = get_connection()
            if not conn:
                lbl_err.text = 'Error de conexión'
                return
            cur = conn.cursor()
            try:
                if es_edicion:
                    cur.execute("""
                        UPDATE suscriptores
                        SET nombre=%s, direccion=%s, municipio=%s, barrio=%s,
                            subcategoria=%s, estrato=%s, estado_suministro=%s
                        WHERE cuenta=%s
                    """, (
                        nombre_val,
                        inputs['direccion'].text.strip() or None,
                        inputs['municipio'].text.strip() or None,
                        inputs['barrio'].text.strip() or None,
                        inputs['subcategoria'].text.strip() or None,
                        inputs['estrato'].text.strip() or None,
                        inputs['estado_suministro'].text.strip() or None,
                        cuenta_editar,
                    ))
                else:
                    cuenta_val = inputs['cuenta'].text.strip()
                    susccodi_val = inputs['susccodi'].text.strip()
                    if not cuenta_val:
                        lbl_err.text = 'El número de cuenta es obligatorio'
                        cur.close(); conn.close()
                        return
                    cur.execute("""
                        INSERT INTO suscriptores
                        (cuenta, susccodi, nombre, direccion, municipio,
                         barrio, subcategoria, estrato, estado_suministro)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        int(cuenta_val),
                        int(susccodi_val) if susccodi_val else None,
                        nombre_val,
                        inputs['direccion'].text.strip() or None,
                        inputs['municipio'].text.strip() or None,
                        inputs['barrio'].text.strip() or None,
                        inputs['subcategoria'].text.strip() or None,
                        inputs['estrato'].text.strip() or None,
                        inputs['estado_suministro'].text.strip() or None,
                    ))
                conn.commit()
                lbl_ok.text = 'Guardado correctamente'
                lbl_err.text = ''
                self.mostrar_todos()
            except Exception as e:
                conn.rollback()
                lbl_err.text = f'Error: {e}'
            finally:
                cur.close()
                conn.close()

        btn_guardar.bind(on_press=guardar)
        btn_cancelar.bind(on_press=popup.dismiss)
        popup.open()

    # ------------------------------------------------------------------
    #  Exportar catastro completo a Excel (hilo → openpyxl)
    # ------------------------------------------------------------------
    def exportar_excel(self):
        overlay.show('Cargando opciones…')

        def _tarea():
            conn = get_connection()
            if not conn:
                Clock.schedule_once(
                    lambda *_: (overlay.hide(),
                                setattr(self, 'mensaje', 'Error de conexión')), 0)
                return
            cur = conn.cursor()
            try:
                cur.execute("""
                    SELECT DISTINCT estrato FROM suscriptores
                    WHERE estrato IS NOT NULL AND estrato != ''
                    ORDER BY CAST(estrato AS UNSIGNED)
                """)
                estratos = [str(r[0]) for r in cur.fetchall()]
                cur.execute("SELECT DISTINCT barrio FROM suscriptores WHERE barrio IS NOT NULL AND barrio != '' ORDER BY barrio")
                barrios = [r[0] for r in cur.fetchall()]
            finally:
                cur.close()
                conn.close()
            Clock.schedule_once(lambda *_: self._popup_config_excel(estratos, barrios), 0)

        threading.Thread(target=_tarea, daemon=True).start()

    def _popup_config_excel(self, estratos, barrios):
        overlay.hide()
        from kivy.uix.togglebutton import ToggleButton

        content = BoxLayout(orientation='vertical', spacing=0)
        with content.canvas.before:
            Color(*CARD)
            r = Rectangle(pos=content.pos, size=content.size)
        content.bind(pos=lambda _, v: setattr(r, 'pos', v))
        content.bind(size=lambda _, v: setattr(r, 'size', v))

        # Cabecera
        hdr = BoxLayout(size_hint_y=None, height=52, padding=[16, 0])
        with hdr.canvas.before:
            Color(*TINTA)
            rh = Rectangle(pos=hdr.pos, size=hdr.size)
        hdr.bind(pos=lambda _, v: setattr(rh, 'pos', v))
        hdr.bind(size=lambda _, v: setattr(rh, 'size', v))
        hdr.add_widget(Label(text='Configurar exportación Excel', bold=True,
                             font_size=14, color=(1, 1, 1, 1),
                             halign='left', valign='middle', text_size=(500, 52)))
        content.add_widget(hdr)

        body = BoxLayout(orientation='vertical', padding=[16, 12], spacing=10)
        content.add_widget(body)

        # ── Agrupación ────────────────────────────────────────────
        body.add_widget(Label(text='Agrupar por:', font_size=11, color=MUTED,
                              size_hint_y=None, height=18,
                              halign='left', text_size=(540, 18)))
        grp_row = BoxLayout(size_hint_y=None, height=36, spacing=8)
        btn_est = ToggleButton(text='Estrato', group='agrup', state='down',
                               size_hint_x=None, width=100, font_size=12,
                               background_normal='', background_down='',
                               background_color=TINTA, color=(1, 1, 1, 1))
        btn_bar = ToggleButton(text='Barrio', group='agrup', state='normal',
                               size_hint_x=None, width=100, font_size=12,
                               background_normal='', background_down='',
                               background_color=STAGE, color=TINTA)

        def _sync_colores(*_):
            btn_est.background_color = TINTA if btn_est.state == 'down' else STAGE
            btn_est.color = (1, 1, 1, 1) if btn_est.state == 'down' else TINTA
            btn_bar.background_color = TINTA if btn_bar.state == 'down' else STAGE
            btn_bar.color = (1, 1, 1, 1) if btn_bar.state == 'down' else TINTA

        btn_est.bind(state=_sync_colores)
        btn_bar.bind(state=_sync_colores)
        grp_row.add_widget(btn_est)
        grp_row.add_widget(btn_bar)
        grp_row.add_widget(Label())
        body.add_widget(grp_row)

        # ── Selector de estratos ──────────────────────────────────
        body.add_widget(Label(text='Estratos a incluir (todos por defecto):',
                              font_size=11, color=MUTED,
                              size_hint_y=None, height=18,
                              halign='left', text_size=(540, 18)))

        est_row = BoxLayout(size_hint_y=None, height=38, spacing=6)
        est_btns = {}
        for est in estratos:
            tb = ToggleButton(text=f'E{est}', state='down',
                              size_hint_x=None, width=52, font_size=11,
                              background_normal='', background_down='',
                              background_color=TINTA, color=(1, 1, 1, 1))
            def _sync_est(inst, val):
                inst.background_color = TINTA if inst.state == 'down' else STAGE
                inst.color = (1, 1, 1, 1) if inst.state == 'down' else TINTA
            tb.bind(state=_sync_est)
            est_btns[est] = tb
            est_row.add_widget(tb)

        # Botones Todos / Ninguno
        def _todos(_):
            for tb in est_btns.values():
                tb.state = 'down'
        def _ninguno(_):
            for tb in est_btns.values():
                tb.state = 'normal'

        btn_todos   = Button(text='Todos', size_hint_x=None, width=60, font_size=10,
                             background_normal='', background_color=STAGE, color=TINTA)
        btn_ninguno = Button(text='Ninguno', size_hint_x=None, width=64, font_size=10,
                             background_normal='', background_color=STAGE, color=TINTA)
        btn_todos.bind(on_press=_todos)
        btn_ninguno.bind(on_press=_ninguno)
        est_row.add_widget(Label(size_hint_x=None, width=10))
        est_row.add_widget(btn_todos)
        est_row.add_widget(btn_ninguno)
        est_row.add_widget(Label())
        body.add_widget(est_row)

        # ── Barrios a incluir ─────────────────────────────────────
        body.add_widget(Label(text='Barrios a incluir (todos por defecto):',
                              font_size=11, color=MUTED,
                              size_hint_y=None, height=18,
                              halign='left', text_size=(540, 18)))

        bar_scroll = ScrollView(size_hint_y=None, height=72)
        bar_grid   = GridLayout(cols=5, size_hint_y=None, spacing=4, padding=[0, 2])
        bar_grid.bind(minimum_height=bar_grid.setter('height'))
        bar_btns = {}
        for bar in barrios:
            tb = ToggleButton(text=bar[:14], state='down',
                              size_hint_y=None, height=30, font_size=10,
                              background_normal='', background_down='',
                              background_color=TINTA, color=(1, 1, 1, 1))
            def _sync_bar(inst, val):
                inst.background_color = TINTA if inst.state == 'down' else STAGE
                inst.color = (1, 1, 1, 1) if inst.state == 'down' else TINTA
            tb.bind(state=_sync_bar)
            bar_btns[bar] = tb
            bar_grid.add_widget(tb)
        bar_scroll.add_widget(bar_grid)

        bar_ctrl = BoxLayout(size_hint_y=None, height=28, spacing=6)
        btn_bar_todos   = Button(text='Todos', size_hint_x=None, width=60, font_size=10,
                                 background_normal='', background_color=STAGE, color=TINTA)
        btn_bar_ninguno = Button(text='Ninguno', size_hint_x=None, width=64, font_size=10,
                                 background_normal='', background_color=STAGE, color=TINTA)
        def _bar_todos(_):
            for tb in bar_btns.values(): tb.state = 'down'
        def _bar_ninguno(_):
            for tb in bar_btns.values(): tb.state = 'normal'
        btn_bar_todos.bind(on_press=_bar_todos)
        btn_bar_ninguno.bind(on_press=_bar_ninguno)
        bar_ctrl.add_widget(btn_bar_todos)
        bar_ctrl.add_widget(btn_bar_ninguno)
        bar_ctrl.add_widget(Label())
        body.add_widget(bar_ctrl)
        body.add_widget(bar_scroll)

        # ── Footer ────────────────────────────────────────────────
        footer = BoxLayout(size_hint_y=None, height=56, spacing=10, padding=[16, 10])
        with footer.canvas.before:
            Color(*STAGE)
            rf = Rectangle(pos=footer.pos, size=footer.size)
        footer.bind(pos=lambda _, v: setattr(rf, 'pos', v))
        footer.bind(size=lambda _, v: setattr(rf, 'size', v))

        popup = Popup(title='', content=content, size_hint=(0.58, None), height=500,
                      background_color=CARD, separator_height=0)

        def _exportar(_):
            agrup = 'estrato' if btn_est.state == 'down' else 'barrio'
            ests_sel = [e for e, tb in est_btns.items() if tb.state == 'down']
            bars_sel = [b for b, tb in bar_btns.items() if tb.state == 'down']
            if not ests_sel:
                return
            popup.dismiss()
            self._ejecutar_exportacion(agrup, ests_sel, bars_sel)

        btn_cancel = Button(text='Cancelar', size_hint=(0.3, None), height=36,
                            background_normal='', background_color=LINE, color=TINTA,
                            font_size=12)
        btn_export = Button(text='↓ Exportar Excel', size_hint=(0.7, None), height=36,
                            background_normal='', background_color=TINTA,
                            color=(1, 1, 1, 1), font_size=12)
        btn_cancel.bind(on_press=popup.dismiss)
        btn_export.bind(on_press=_exportar)
        footer.add_widget(btn_cancel)
        footer.add_widget(btn_export)
        content.add_widget(footer)
        popup.open()

    def _ejecutar_exportacion(self, agrup, estratos_sel, barrios_sel):
        overlay.show('Generando Excel…')

        def _tarea():
            rows, error = [], None
            conn = get_connection()
            if not conn:
                Clock.schedule_once(
                    lambda *_: (overlay.hide(),
                                setattr(self, 'mensaje', 'Error de conexión')), 0)
                return
            cur = conn.cursor()
            try:
                ph_est = ', '.join(['%s'] * len(estratos_sel))
                ph_bar = ', '.join(['%s'] * len(barrios_sel)) if barrios_sel else None
                if agrup == 'estrato':
                    order = 'CAST(estrato AS UNSIGNED), barrio, nombre'
                else:
                    order = 'barrio, CAST(estrato AS UNSIGNED), nombre'
                if ph_bar:
                    sql = f"""
                        SELECT nombre, estrato, direccion, barrio
                        FROM suscriptores
                        WHERE estrato IN ({ph_est}) AND barrio IN ({ph_bar})
                        ORDER BY {order}
                    """
                    cur.execute(sql, estratos_sel + barrios_sel)
                else:
                    sql = f"""
                        SELECT nombre, estrato, direccion, barrio
                        FROM suscriptores
                        WHERE estrato IN ({ph_est})
                        ORDER BY {order}
                    """
                    cur.execute(sql, estratos_sel)
                rows = cur.fetchall()
            except Exception as e:
                error = str(e)
            finally:
                cur.close()
                conn.close()
            Clock.schedule_once(lambda *_: _aplicar(rows, error), 0)

        def _aplicar(rows, error):
            overlay.hide()
            if error:
                self.mensaje = f'Error al consultar: {error}'
                return
            if not rows:
                self.mensaje = 'No hay suscriptores con esos filtros'
                return
            try:
                ruta = _generar_xlsx(rows, agrup)
                self.mensaje = f'Excel guardado — {len(rows):,} suscriptores'
                os.startfile(ruta)
            except Exception as e:
                self.mensaje = f'Error al generar Excel: {e}'

        threading.Thread(target=_tarea, daemon=True).start()


def _generar_xlsx(rows, agrupacion='estrato'):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    TINTA_HEX = '1B1512'
    CREMA_HEX = 'F1EAE1'
    STAGE_HEX = 'E8DDD4'
    WHITE_HEX = 'FFFFFF'
    RESUMEN_HEX = 'F5F0EA'   # fondo fila de total

    GROUP_COLORS = ['C8E6C9', 'BBDEFB', 'FFF9C4', 'FFE0B2', 'E1BEE7', 'F8BBD0',
                    'B2DFDB', 'FFCCBC', 'D1C4E9', 'B3E5FC']

    def _fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def _border():
        return Border(bottom=Side(style='thin', color='D0C8BE'))

    def _group_fill(key, color_map):
        if key not in color_map:
            color_map[key] = GROUP_COLORS[len(color_map) % len(GROUP_COLORS)]
        return _fill(color_map[key])

    # Decidir clave de agrupación
    # rows: (nombre, estrato, direccion, barrio)
    KEY_IDX = 1 if agrupacion == 'estrato' else 3   # índice en la tupla

    wb = Workbook()
    ws = wb.active
    ws.title = 'Catastro'

    agrup_label = 'Estrato' if agrupacion == 'estrato' else 'Barrio'

    # ── Título ────────────────────────────────────────────────
    ws.merge_cells('A1:D1')
    t = ws['A1']
    t.value = (f'Catastro de Suscriptores — INGESAM  ·  {date.today():%d/%m/%Y}'
               f'   (agrupado por {agrup_label})')
    t.font = Font(name='Calibri', bold=True, size=13, color=WHITE_HEX)
    t.fill = _fill(TINTA_HEX)
    t.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

    # ── Encabezados ───────────────────────────────────────────
    for col, h in enumerate(['Nombre', 'Estrato', 'Dirección', 'Barrio'], start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(name='Calibri', bold=True, size=11, color=CREMA_HEX)
        c.fill = _fill(TINTA_HEX)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 22

    # ── Datos agrupados ───────────────────────────────────────
    color_map   = {}
    current_key = object()   # sentinel
    group_start = 3
    group_count = 0
    data_row    = 3
    row_idx     = 0

    def _escribir_resumen(key, count):
        nonlocal data_row
        ws.merge_cells(f'A{data_row}:C{data_row}')
        lbl = 'Estrato' if agrupacion == 'estrato' else 'Barrio'
        c = ws.cell(row=data_row, column=1,
                    value=f'    Total {lbl} {key or "Sin clasificar"}: {count:,} suscriptores')
        c.font  = Font(name='Calibri', bold=True, size=9, color='5C524C')
        c.fill  = _fill(RESUMEN_HEX)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        # celda D vacía con el mismo fondo
        d = ws.cell(row=data_row, column=4)
        d.fill = _fill(RESUMEN_HEX)
        ws.row_dimensions[data_row].height = 15
        data_row += 1

    for nombre, estrato, direccion, barrio in rows:
        key = (estrato if agrupacion == 'estrato' else barrio) or ''

        if key != current_key:
            # Cerrar grupo anterior con fila de resumen
            if current_key is not object() and group_count > 0:
                _escribir_resumen(current_key, group_count)

            current_key = key
            group_count = 0
            row_idx     = 0

            # Separador de grupo
            ws.merge_cells(f'A{data_row}:D{data_row}')
            label = (f'  Estrato {key or "Sin estrato"}' if agrupacion == 'estrato'
                     else f'  Barrio: {key or "Sin barrio"}')
            sep = ws.cell(row=data_row, column=1, value=label)
            sep.font = Font(name='Calibri', bold=True, size=10, color=TINTA_HEX)
            sep.fill = _group_fill(key, color_map)
            sep.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[data_row].height = 18
            data_row += 1

        # Fila de dato
        bg = _fill(WHITE_HEX if row_idx % 2 == 0 else STAGE_HEX)
        for col, val in enumerate([nombre or '', estrato or '',
                                   direccion or '', barrio or ''], start=1):
            c = ws.cell(row=data_row, column=col, value=val)
            c.font      = Font(name='Calibri', size=10, color=TINTA_HEX)
            c.fill      = bg
            c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            c.border    = _border()
        ws.row_dimensions[data_row].height = 16
        data_row  += 1
        row_idx   += 1
        group_count += 1

    # Resumen del último grupo
    if group_count > 0:
        _escribir_resumen(current_key, group_count)

    # ── Anchos ────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 22
    ws.freeze_panes = 'A3'

    nombre_archivo = f'catastro_{date.today():%Y-%m-%d}.xlsx'
    ruta = os.path.join(os.path.expanduser('~'), 'Documents', nombre_archivo)
    wb.save(ruta)
    return ruta
